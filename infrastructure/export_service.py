import os
import base64
import datetime
import calendar
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from infrastructure.logging_service import get_module_logger
from domain.cost import CostType
from domain.quote_validator import QuoteValidator

logger = get_module_logger("ExportService", "export_project.log")


class ExportService:
    """
    Export Excel basé sur un template avec des placeholders pré-formatés.
    - Lignes 21 à 30 : lignes de quantité avec PART_REF (Col B), QTY_REF (Col I), PU_REF (Col K)
    - Autres placeholders globaux: DEVIS_REF, DATE_REF, VALIDITY_REF, COMMENT_REF
    
    Utilise QuoteNumberingService pour gérer les numéros de quote persistants.
    """

    def __init__(self, db=None):
        """Initialize avec optionnel DB pour numbering service."""
        self.db = db
        self.numbering_service = None
        if db:
            from infrastructure.quote_numbering_service import QuoteNumberingService
            self.numbering_service = QuoteNumberingService(db)

    def get_devis_reference(self, project=None, sub_version: int = None):
        """
        Génère la référence du devis avec numérotation persistante.
        Format: {prefix}{date}_{counter:03d}-{sub_version}
        Exemple: OD260202_001-1
        
        Logique:
        - Si la quote a déjà un numéro (dans export_history), le réutiliser
        - Sinon, assigner un nouveau numéro avec le compteur du jour
        - La sous-version s'incrémente à chaque export, peu importe le jour
        
        Args:
            project: Project object (for compatibility, can be None if using numbering_service)
            sub_version: Sub-version number. If None, uses history count for that project
        """
        
        # Si numbering service disponible, l'utiliser (meilleur système)
        if self.numbering_service and project and hasattr(project, 'export_history'):
            # Chercher si cette quote a déjà un numéro assigné dans son historique
            existing_ref = None
            for export_entry in project.export_history:
                devis_ref = export_entry.get('devis_ref') or export_entry.get('reference')
                if devis_ref and devis_ref.startswith('OD'):
                    # Extraire le numéro (sans la sous-version)
                    # Format: OD260202_001-1 → on veut OD260202_001
                    if '-' in devis_ref:
                        existing_ref = devis_ref.rsplit('-', 1)[0]
                    else:
                        existing_ref = devis_ref
                    break
            
            # Si un numéro existe, le réutiliser
            if existing_ref:
                quote_num = existing_ref
            else:
                # Sinon, obtenir un nouveau numéro (incrémente le compteur du jour)
                quote_num, counter = self.numbering_service.get_next_quote_number("OD")
            
            # Déterminer la sous-version (nombre d'exports précédents de cette quote + 1)
            if sub_version is None:
                sub_version = len(project.export_history) + 1
            
            return f"{quote_num}-{sub_version}"
        
        # Fallback ancien système (compatible)
        today = datetime.date.today()
        base_ref = f"OD{today.strftime('%y%m%d')}_001"
        
        if project and hasattr(project, 'export_history'):
            # Chercher si cette quote a déjà un numéro
            existing_ref = None
            for export_entry in project.export_history:
                devis_ref = export_entry.get('devis_ref') or export_entry.get('reference')
                if devis_ref and devis_ref.startswith('OD'):
                    if '-' in devis_ref:
                        existing_ref = devis_ref.rsplit('-', 1)[0]
                    else:
                        existing_ref = devis_ref
                    break
            
            if existing_ref:
                base_ref = existing_ref
            
            sub_version = len(project.export_history) + 1
            return f"{base_ref}-{sub_version}"
            
        return base_ref

    def get_default_filename(self, project, devis_ref: str = None):
        """Génère le nom de fichier par défaut : DEVIS_REF-PART_REF-xQTYmin-xQTYmax.xlsx"""
        if devis_ref is None:
            devis_ref = self.get_devis_reference(project)
        part_ref = self._get_part_reference(project)
        # Supprimer les espaces et caractères spéciaux du part_ref pour le nom de fichier
        clean_part_ref = "".join(c for c in part_ref if c.isalnum() or c in ('-', '_')).strip()
        
        quantities = sorted(project.sale_quantities)
        if not quantities:
            return f"{devis_ref}-{clean_part_ref}.xlsx"
        
        q_min = self._format_qty(quantities[0])
        q_max = self._format_qty(quantities[-1])
        
        if len(quantities) == 1:
            return f"{devis_ref}-{clean_part_ref}-x{q_min}.xlsx"
        else:
            return f"{devis_ref}-{clean_part_ref}-x{q_min}-x{q_max}.xlsx"

    def _format_qty(self, qty):
        """Formate la quantité avec x100, x1k, x10k etc. avec arrondis intelligents."""
        if qty <= 0: return "0"
        
        # Gestion des arrondis si proche de 100 ou 1000
        # (Si la différence est < 2%, on arrondit pour le nom de fichier)
        for base in [100, 1000, 10000]:
            if abs(qty - base) / base < 0.02:
                qty = base
                break

        if qty >= 1000000:
            val = qty / 1000000
            return f"{val:g}m"
        elif qty >= 1000:
            val = qty / 1000
            return f"{val:g}k"
        else:
            return str(int(qty))

    # =========================
    # PUBLIC
    # =========================
    def export_excel(self, project, template_path, output_path, project_save_path=None, devis_ref=None):
        try:
            wb = load_workbook(template_path)
            ws = wb.active

            # 0. Generate devis reference
            ref_date = datetime.date.today()
            
            # IMPORTANT: Compter les exports précédents AVANT d'ajouter le nouveau
            # pour que la sous-version soit correcte
            if not hasattr(project, 'export_history'):
                project.export_history = []
            
            if devis_ref is None:
                devis_ref = self.get_devis_reference(project)

            # Keep a fresh diagnostic snapshot during exports too.
            project.validation_report = QuoteValidator.validate(project)
            
            # THE CRITICAL STEP: Add to history NOW so placeholders use THE SAME reference
            project.export_history.append({
                "devis_ref": devis_ref,
                "date": ref_date.strftime("%d/%m/%Y"),
                "time": datetime.datetime.now().strftime("%H:%M"),
                "version_index": getattr(project, 'current_version_index', 1),
            })

            # Placeholder replacement will now call get_devis_reference again, 
            # but daily_count will have increased, so we MUST use the devis_ref we just generated.
            self._current_export_ref = devis_ref

            # 1. Trier les quantités par ordre croissant
            quantities = sorted(project.sale_quantities)
            qty_count = len(quantities)
            qty_processed = 0

            # 2. Préparer le commentaire global dans l'ordre des opérations:
            # commentaires d'opérations + commentaires méthode des coûts outillage.
            global_comment = self._build_global_comment(project)
            tooling_lines = self._get_tooling_lines(project)

            logger.info(f"Export {qty_count} quantités vers le template.")

            # 3. Identifier les lignes de placeholders et les remplir
            rows_to_hide = []
            
            for row_idx in range(1, ws.max_row + 1):
                # La ligne 33 doit rester statique
                if row_idx == 33:
                    continue

                is_qty_row = False
                
                # Vérifier si c'est une ligne de quantité (PART_REF en colonne B)
                # STRICTEMENT limité aux lignes 21 à 30
                if 21 <= row_idx <= 30:
                    cell_b = ws.cell(row=row_idx, column=2)
                    if cell_b.value == "PART_REF":
                        if qty_processed < qty_count:
                            qty = quantities[qty_processed]
                            self._fill_qty_row(ws, row_idx, project, qty, global_comment)
                            qty_processed += 1
                            is_qty_row = True
                        else:
                            # Effacer les placeholders de la ligne avant de la masquer
                            self._clear_row_placeholders(ws, row_idx)
                            rows_to_hide.append(row_idx)
                            continue

                # 4. Remplacer les placeholders globaux sur toute la ligne
                if not is_qty_row:
                    self._replace_global_placeholders(ws, row_idx, project, global_comment)

            # 5. Masquer les lignes en trop (au lieu de les supprimer pour préserver les fusions)
            for row_idx in rows_to_hide:
                ws.row_dimensions[row_idx].hidden = True

            # 5.b Gestion des lignes outillage (TOOL_REF / QTY_REF / PTOOL_REF)
            self._fill_tooling_rows(ws, tooling_lines, project, global_comment)
            # 5.c Sécurise COMMENT_REF (certaines manipulations de lignes peuvent l'effacer)
            self._fill_comment_placeholders(ws, global_comment)

            # 7. Ajouter l'onglet SERIE si le mode série est actif
            if getattr(project, 'serie_data', None) is not None:
                self._add_serie_sheet(wb, project.serie_data)

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)

            logger.info(f"Export Excel réussi → {output_path}")

            # Embed the generated XLSX bytes in the last export_history entry
            try:
                with open(output_path, 'rb') as f:
                    xlsx_bytes = f.read()
                xlsx_filename = os.path.basename(output_path)
                if project.export_history:
                    last_entry = project.export_history[-1]
                    last_entry['xlsx_filename'] = xlsx_filename
                    last_entry['xlsx_data_b64'] = base64.b64encode(xlsx_bytes).decode('ascii')
                    last_entry['_xlsx_path'] = f"documents/exports/{xlsx_filename}"
            except Exception as e:
                logger.warning(f"Impossible d'embarquer le XLSX dans le projet: {e}")

            # 6. Save project to persist history if path provided
            if project_save_path:
                from infrastructure.persistence import PersistenceService
                PersistenceService.save_project(project, project_save_path)
                logger.info(f"Projet mis à jour avec l'historique d'export → {project_save_path}")

        except PermissionError:
            msg = f"Impossible d'enregistrer le fichier '{os.path.basename(output_path)}'. Vérifiez qu'il n'est pas déjà ouvert dans Excel."
            logger.error(msg)
            raise Exception(msg)
        except Exception as e:
            logger.error("Erreur export Excel", exc_info=True)
            raise

    # =========================
    # PRIVATE
    # =========================
    def _get_part_reference(self, project):
        return getattr(project, "reference", "")

    def _fill_qty_row(self, ws, row_idx, project, qty, global_comment):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if not isinstance(val, str):
                continue
            token = val.strip()
                
            if token == "PART_REF":
                cell.value = self._get_part_reference(project)
            elif token == "QTY_REF":
                cell.value = qty
            elif token == "PU_REF":
                pu = project.total_price(qty)
                cell.value = pu
                cell.number_format = '€ #,##0.00'
            else:
                self._replace_cell_placeholder(cell, project, global_comment)

    def _replace_global_placeholders(self, ws, row_idx, project, global_comment):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            self._replace_cell_placeholder(cell, project, global_comment)

    def _replace_cell_placeholder(self, cell, project, global_comment):
        val = cell.value
        if not isinstance(val, str):
            return
        token = val.strip()

        # Determine reference date (used for both DATE_REF and VALIDITY_REF)
        # Always use today's date at the moment of export
        ref_date = datetime.date.today()

        if token == "DEVIS_REF":
            # Use the reference pre-calculated for this export session
            if hasattr(self, '_current_export_ref'):
                cell.value = self._current_export_ref
            else:
                cell.value = self.get_devis_reference(project)
        elif token == "PART_REF":
            cell.value = self._get_part_reference(project)
        elif token == "CUSTOMER_NAME":
            cell.value = getattr(project, "client", "")
        elif token == "DATE_REF":
            cell.value = ref_date.strftime("%d/%m/%Y")
        elif token == "VALIDITY_REF":
            # Add exactly 1 month to ref_date
            month = ref_date.month - 1 + 1
            year = ref_date.year + month // 12
            month = month % 12 + 1
            day = min(ref_date.day, calendar.monthrange(year, month)[1])
            date_validity = datetime.date(year, month, day)
            cell.value = date_validity.strftime("%d/%m/%Y")
        elif token == "COMMENT_REF":
            cell.value = global_comment
        elif token == "IMAGE_PREVIEW":
            # injecter l'image preview dans le placeholder (zones déjà mergées dans le template)
            project_preview = getattr(project, 'preview_image', None)
            if project_preview and getattr(project_preview, 'data', None):
                try:
                    from openpyxl.drawing.image import Image as OpenpyxlImage
                    from PIL import Image as PilImage
                    import io
                    import base64

                    preview_data = base64.b64decode(project_preview.data)
                    preview_img = PilImage.open(io.BytesIO(preview_data))

                    # Taille de la vignette pour le placeholder (hauteur ligne 210px)
                    max_height = 210
                    max_width = 300
                    if hasattr(PilImage, 'Resampling'):
                        resample = PilImage.Resampling.LANCZOS
                    else:
                        resample = PilImage.ANTIALIAS
                    preview_img.thumbnail((max_width, max_height), resample)

                    image_stream = io.BytesIO()
                    preview_img.save(image_stream, format='PNG')
                    image_stream.seek(0)

                    img = OpenpyxlImage(image_stream)
                    ws = cell.parent
                    
                    # Placer l'image dans la cellule (les zones sont déjà mergées dans le template)
                    img.anchor = cell.coordinate
                    ws.add_image(img)
                    
                    # Ajuster la hauteur de la ligne pour accueillir l'image
                    ws.row_dimensions[cell.row].height = 210
                    
                    # Effacer le texte placeholder
                    cell.value = ""
                except Exception as e:
                    logger.error(f"Erreur lors de l'insertion d'image: {e}", exc_info=True)
                    cell.value = "Preview non disponible"
            else:
                cell.value = "Preview non fournie"

    def _clear_row_placeholders(self, ws, row_idx):
        """Efface les placeholders d'une ligne avant de la masquer."""
        placeholders = [
            "PART_REF", "QTY_REF", "PU_REF",
            "TOOL_REF", "PTOOL_REF",
            "DEVIS_REF", "DATE_REF", "VALIDITY_REF", "CUSTOMER_NAME"
        ]
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str) and cell.value.strip() in placeholders:
                cell.value = ""

    def _get_tooling_lines(self, project):
        lines = []
        for op in getattr(project, "operations", []):
            for cost in getattr(op, "costs", {}).values():
                if getattr(cost, "cost_type", None) == CostType.TOOLING:
                    lines.append(cost)
        return lines

    def _find_rows_with_placeholder(self, ws, placeholder):
        rows = []
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, str) and val.strip() == placeholder:
                    rows.append(row_idx)
                    break
        return rows

    def _clone_row_format(self, ws, source_row, target_row):
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col_idx)
            dst = ws.cell(row=target_row, column=col_idx)
            dst.value = src.value
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
        src_dim = ws.row_dimensions[source_row]
        ws.row_dimensions[target_row].height = src_dim.height
        ws.row_dimensions[target_row].hidden = src_dim.hidden

    def _fill_tool_row(self, ws, row_idx, tool_cost, project, global_comment):
        price = 0.0
        if getattr(tool_cost, "pricing", None):
            # Outillage exporté en ligne unitaire x1: on utilise le montant de lot.
            price = float(getattr(tool_cost.pricing, "fixed_price", 0.0) or 0.0)

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if not isinstance(val, str):
                continue
            token = val.strip()
            if token == "TOOL_REF":
                cell.value = tool_cost.name
            elif token == "QTY_REF":
                cell.value = 1
            elif token == "PTOOL_REF":
                cell.value = price
                cell.number_format = '€ #,##0.00'
            else:
                self._replace_cell_placeholder(cell, project, global_comment)
        ws.row_dimensions[row_idx].hidden = False

    def _fill_tooling_rows(self, ws, tooling_lines, project, global_comment):
        tool_rows = self._find_rows_with_placeholder(ws, "TOOL_REF")
        if not tool_rows:
            return

        if not tooling_lines:
            for row_idx in tool_rows:
                self._clear_row_placeholders(ws, row_idx)
                ws.row_dimensions[row_idx].hidden = True
            return

        n_tools = len(tooling_lines)
        n_template = len(tool_rows)

        # Si plus d'outillages que de lignes template → cloner la dernière
        if n_tools > n_template:
            extra_needed = n_tools - n_template
            last_row = tool_rows[-1]
            ws.insert_rows(last_row + 1, amount=extra_needed)
            for i in range(extra_needed):
                self._clone_row_format(ws, last_row, last_row + 1 + i)
            tool_rows = tool_rows + [last_row + 1 + i for i in range(extra_needed)]

        for idx, tool_cost in enumerate(tooling_lines):
            self._fill_tool_row(ws, tool_rows[idx], tool_cost, project, global_comment)

        # Masquer les lignes template non utilisées
        for row_idx in tool_rows[n_tools:]:
            self._clear_row_placeholders(ws, row_idx)
            ws.row_dimensions[row_idx].hidden = True

    def _fill_comment_placeholders(self, ws, global_comment):
        replaced = 0
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, str) and cell.value.strip() == "COMMENT_REF":
                    cell.value = global_comment
                    replaced += 1
        logger.info(f"COMMENT_REF replaced count: {replaced}")

    def _build_global_comment(self, project):
        op_lines = []
        tooling_comments = []
        for op in getattr(project, "operations", []):
            if getattr(op, "comment", None) and op.comment.strip():
                op_lines.append(op.comment.strip())
            for cost in getattr(op, "costs", {}).values():
                if getattr(cost, "cost_type", None) == CostType.TOOLING:
                    if getattr(cost, "client_comment", None) and cost.client_comment.strip():
                        tooling_comments.append(cost.client_comment.strip())

        lines = op_lines + tooling_comments

        text = "\n".join(lines) if lines else "-"
        logger.info(f"Global comment lines: {len(lines)}")
        return text

    def _add_serie_sheet(self, wb, sd):
        """Ajoute un onglet 'SERIE' au workbook Excel avec le chiffrage gros volumes."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        FONT = "Arial"
        H_BG  = "1F4E79"   # bleu section
        H_FG  = "FFFFFF"
        TOT_BG = "2E75B6"
        RES_BG = "FFFF99"
        PV_BG  = "375623"

        def hdr(cell, text):
            cell.value = text
            cell.font = Font(name=FONT, bold=True, size=10, color=H_FG)
            cell.fill = PatternFill("solid", fgColor=H_BG)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        def inp(cell, value, fmt="#,##0.00"):
            cell.value = value
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = fmt

        def res(cell, value, fmt="#,##0.0000"):
            cell.value = value
            cell.font = Font(name=FONT, bold=True, size=9)
            cell.fill = PatternFill("solid", fgColor=RES_BG)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = fmt

        def lbl(cell, text, bold=False):
            cell.value = text
            cell.font = Font(name=FONT, bold=bold, size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        def tot(cell, value, fmt="#,##0.00"):
            cell.value = value
            cell.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=TOT_BG)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = fmt

        ws = wb.create_sheet("SERIE")

        # Column widths
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 14

        r = 1
        ws.merge_cells(f"A{r}:E{r}")
        c = ws[f"A{r}"]
        c.value = "CHIFFRAGE PRODUCTION SERIE – Gros Volumes"
        c.font = Font(name=FONT, bold=True, size=13, color=H_FG)
        c.fill = PatternFill("solid", fgColor=H_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1

        # ---- HYPOTHESES ----
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], "HYPOTHESES GENERALES")
        ws.row_dimensions[r].height = 18
        r += 1

        hyp_rows = [
            ("Volume annuel cible", sd.annual_volume, "#,##0", "pcs/an"),
            ("Jours ouvrés / an", sd.working_days_per_year, "#,##0", "j/an"),
            ("Equipes / jour", sd.shifts_per_day, "0", ""),
            ("Heures / équipe", sd.hours_per_shift, "0.0", "h"),
            ("TRS", sd.trs, "0.0%", ""),
            ("Temps de cycle goulot (calculé)", sd.get_target_cycle_time_s(), "0.00", "s/pcs"),
            ("TH MO Production", sd.mo_production_rate, "#,##0.00", "€/h"),
            ("TH MO Qualité", sd.mo_quality_rate, "#,##0.00", "€/h"),
            ("Coef. structure overhead", sd.overhead_coef, "0%", ""),
            ("Capacité réelle / an", sd.real_capacity_per_year(), "#,##0", "pcs/an [calculé]"),
            ("Taux de charge", sd.load_rate(), "0.0%", ""),
        ]
        for label, val, fmt, unit in hyp_rows:
            lbl(ws[f"A{r}"], label)
            inp(ws[f"B{r}"], val, fmt)
            lbl(ws[f"C{r}"], unit)
            r += 1

        # ---- CAPEX ----
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], f"CAPEX – INVESTISSEMENTS  |  Amort. = durée programme ({sd.program_lifetime_years} ans)  |  Marge globale {sd.capex_global_margin*100:.0f}%")
        ws.row_dimensions[r].height = 18
        r += 1
        for col, text in zip(["A","B","C","D","E"],
                              ["Désignation","Coût €","Résiduel €","Marge","Prix/pc €"]):
            c = ws[f"{col}{r}"]
            c.value = text
            c.font = Font(name=FONT, bold=True, size=9, color="9C6500")
            c.fill = PatternFill("solid", fgColor="FFEB9C")
            c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
        years = sd.program_lifetime_years if sd.program_lifetime_years > 0 else 1
        capex_total_pc = 0.0
        for item in sd.capex_items:
            amort_an = (item.cost - item.residual_value) / years
            price_pc = (amort_an / sd.annual_volume * (1 + item.margin_rate)) if sd.annual_volume > 0 else 0
            capex_total_pc += price_pc
            lbl(ws[f"A{r}"], item.name)
            inp(ws[f"B{r}"], item.cost, "#,##0")
            inp(ws[f"C{r}"], item.residual_value, "#,##0")
            inp(ws[f"D{r}"], item.margin_rate, "0%")
            inp(ws[f"E{r}"], price_pc, "#,##0.0000")
            r += 1
        tot(ws[f"A{r}"], "TOTAL CAPEX / pièce")
        tot(ws[f"E{r}"], sd.capex_price_per_piece(), "#,##0.0000")
        r += 1

        # ---- TOOLING ----
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], "TOOLING / OUTILLAGES")
        ws.row_dimensions[r].height = 18
        r += 1
        for col, text in zip(["A","B","C","D","E"],
                              ["Désignation","Coût €","Durée vie pcs","Marge","Prix/pc €"]):
            c = ws[f"{col}{r}"]
            c.value = text
            c.font = Font(name=FONT, bold=True, size=9, color="9C6500")
            c.fill = PatternFill("solid", fgColor="FFEB9C")
            c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
        for item in sd.tooling_items:
            cost_pc  = (item.cost / item.lifetime_pieces) if item.lifetime_pieces > 0 else 0
            price_pc = cost_pc * (1 + item.margin_rate)
            lbl(ws[f"A{r}"], item.name)
            inp(ws[f"B{r}"], item.cost, "#,##0")
            inp(ws[f"C{r}"], item.lifetime_pieces, "#,##0")
            inp(ws[f"D{r}"], item.margin_rate, "0%")
            inp(ws[f"E{r}"], price_pc, "#,##0.0000")
            r += 1
        tot(ws[f"A{r}"], "TOTAL TOOLING / pièce")
        tot(ws[f"E{r}"], sd.tooling_price_per_piece(), "#,##0.0000")
        r += 1

        # ---- SETUP ----
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], "SETUP / DEMARRAGE SERIE")
        ws.row_dimensions[r].height = 18
        r += 1
        for label, val, fmt in [
            ("Temps montage outillage", sd.tooling_setup_time_h, "0.00"),
            ("Validation SOP / 1er art.", sd.sop_validation_time_h, "0.00"),
            ("Taille de lot / campagne", sd.lot_size, "#,##0"),
            ("Campagnes / an", sd.campaigns_per_year(), "0.0"),
            ("Coût setup / campagne €", sd.setup_cost_per_campaign(), "#,##0.00"),
            ("Coût setup total / an €", sd.setup_cost_per_year(), "#,##0.00"),
            ("Setup amorti / pièce (prix)", sd.setup_price_per_piece(), "#,##0.0000"),
        ]:
            lbl(ws[f"A{r}"], label)
            inp(ws[f"B{r}"], val, fmt)
            r += 1

        # ---- CONTROLE ----
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], f"CONTROLE DE PRODUCTION (mode : {sd.control_mode})")
        ws.row_dimensions[r].height = 18
        r += 1
        for label, val, fmt in [
            ("Fréquence SPC (1/N pcs)", sd.spc_frequency, "#,##0"),
            ("Temps SPC / pièce mesurée (min)", sd.spc_time_per_piece_min, "0.00"),
            ("Coût SPC / pièce produite €", sd.spc_cost_per_piece(), "#,##0.0000"),
            ("Temps contrôle 100% (s/pcs)", sd.control_100pct_time_s, "0.00"),
            ("Coût contrôle 100% / pièce €", sd.control_100pct_cost_per_piece(), "#,##0.0000"),
            ("Coût contrôle retenu / pièce €", sd.control_cost_per_piece(), "#,##0.0000"),
        ]:
            lbl(ws[f"A{r}"], label)
            inp(ws[f"B{r}"], val, fmt)
            r += 1

        # ---- SYNTHESE ----
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], "SYNTHESE COUT DE REVIENT")
        ws.row_dimensions[r].height = 18
        r += 1
        for col, text in zip(["A","B","C","D"],
                              ["Poste de coût","Coût brut €/pcs","Prix client €/pcs","% du total"]):
            c = ws[f"{col}{r}"]
            c.value = text
            c.font = Font(name=FONT, bold=True, size=9, color="9C6500")
            c.fill = PatternFill("solid", fgColor="FFEB9C")
            c.alignment = Alignment(horizontal="center", vertical="center")
        r += 1

        pv = sd.selling_price_per_piece()
        syn_items = [
            ("MO directe production", sd.mo_cost_per_piece(), sd.mo_cost_per_piece()),
            ("CAPEX amortissement",   sd.capex_cost_per_piece(), sd.capex_price_per_piece()),
            ("Tooling",               sd.tooling_cost_per_piece(), sd.tooling_price_per_piece()),
            ("Setup",                 sd.setup_cost_per_piece(), sd.setup_price_per_piece()),
            ("Contrôle qualité",      sd.control_cost_per_piece(), sd.control_cost_per_piece()),
            ("Matières premières",    sd.material_cost_per_piece, sd.material_cost_per_piece * (1 + sd.material_margin)),
            ("Logistique / emballage",sd.logistics_cost_per_piece, sd.logistics_cost_per_piece * (1 + sd.logistics_margin)),
        ]
        for label, cost, price in syn_items:
            pct = (price / pv) if pv > 0 else 0
            lbl(ws[f"A{r}"], label)
            inp(ws[f"B{r}"], cost, "#,##0.0000")
            inp(ws[f"C{r}"], price, "#,##0.0000")
            inp(ws[f"D{r}"], pct, "0.0%")
            r += 1

        # Sous-total
        tot(ws[f"A{r}"], "Sous-total (avant marge ciale)")
        tot(ws[f"B{r}"], sd.total_cost_per_piece(), "#,##0.0000")
        tot(ws[f"C{r}"], sd.subtotal_with_item_margins(), "#,##0.0000")
        r += 1
        r += 1

        # Prix de vente final
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = f"PRIX DE VENTE / PIECE  (marge ciale {sd.global_commercial_margin*100:.0f}%)"
        c.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=PV_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        c2 = ws[f"C{r}"]
        c2.value = pv
        c2.font = Font(name=FONT, bold=True, size=14, color="FFFFFF")
        c2.fill = PatternFill("solid", fgColor=PV_BG)
        c2.number_format = "#,##0.0000"
        c2.alignment = Alignment(horizontal="right", vertical="center")
        r += 1

        # LOP
        r += 1
        ws.merge_cells(f"A{r}:E{r}")
        hdr(ws[f"A{r}"], f"LIFE OF PROGRAM  —  {sd.program_lifetime_years} ans  |  {sd.total_program_volume():,} pcs totales")
        ws.row_dimensions[r].height = 18
        r += 1
        lop_rows = [
            ("CA total programme", sd.total_program_revenue(), "#,##0"),
            ("Coût total programme", sd.total_program_cost(), "#,##0"),
            ("  dont CAPEX net (fixe, indép. durée)", sd.total_capex_net_investment(), "#,##0"),
            ("  dont Tooling (fixe, indép. durée)", sd.total_tooling_investment(), "#,##0"),
            ("  dont MO + setup + ctrl + mat. (variable)", sd.total_variable_program_cost(), "#,##0"),
        ]
        for label, val, fmt in lop_rows:
            lbl(ws[f"A{r}"], label)
            inp(ws[f"B{r}"], val, fmt)
            lbl(ws[f"C{r}"], "€")
            r += 1

        # CA annuel
        r += 1
        lbl(ws[f"A{r}"], "Chiffre d'affaires annuel estimé")
        c_ca = ws[f"C{r}"]
        c_ca.value = sd.annual_revenue()
        c_ca.font = Font(name=FONT, bold=True, size=11, color="375623")
        c_ca.number_format = "#,##0"
        c_ca.alignment = Alignment(horizontal="right", vertical="center")
        ws[f"D{r}"].value = "€ / an"
        ws[f"D{r}"].font = Font(name=FONT, size=9, color="375623")

    def export_fabrication_quality(self, project, output_path):
        """
        Export Fabrication/Qualité - Document Excel avec la trame de fabrication hiérarchique
        et les informations de temps et commentaires pour chaque opération.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Fabrication_Qualité"

            # Styles
            header_font = Font(bold=True, size=12)
            subheader_font = Font(bold=True, size=10)
            normal_font = Font(size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            header_fill = PatternFill(start_color="FFE6E6FA", end_color="FFE6E6FA", fill_type="solid")

            # En-têtes
            headers = ["Opération", "Typologie", "Temps fixe (h)", "Temps/pièce (h)",
                      "Commentaire Chiffrage", "Commentaire Méthode", "Preview"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                cell.fill = header_fill

            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 30  # Opération
            ws.column_dimensions['B'].width = 20  # Typologie
            ws.column_dimensions['C'].width = 15  # Temps fixe
            ws.column_dimensions['D'].width = 15  # Temps/pièce
            ws.column_dimensions['E'].width = 40  # Commentaire Chiffrage
            ws.column_dimensions['F'].width = 40  # Commentaire Méthode
            ws.column_dimensions['G'].width = 20  # Preview (vignette image ou libellé)

            row = 2

            # Informations du projet en en-tête
            ws.cell(row=row, column=1, value=f"Nom du projet : {getattr(project, 'name', 'N/A')}")
            ws.cell(row=row, column=2, value=f"Référence projet : {getattr(project, 'reference', 'N/A')}")
            row += 1

            # Référence du plan / documents
            drawing_name = getattr(project, 'drawing_filename', None) or 'N/A'
            ws.cell(row=row, column=1, value=f"Référence plan : {drawing_name}")
            ws.cell(row=row, column=2, value=f"Client : {getattr(project, 'client', 'N/A')}")
            row += 1

            # Preview image (si disponible)
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from PIL import Image as PilImage
            import io
            import base64

            if getattr(project, 'preview_image', None) is not None and getattr(project.preview_image, 'data', None):
                try:
                    preview_data = base64.b64decode(project.preview_image.data)
                    preview_img = PilImage.open(io.BytesIO(preview_data))
                    # redimensionner pour éviter une insertion trop grande
                    max_width, max_height = 350, 250
                    preview_img.thumbnail((max_width, max_height), PilImage.Resampling.LANCZOS if hasattr(PilImage, 'Resampling') else PilImage.ANTIALIAS)
                    img_io = io.BytesIO()
                    preview_img.save(img_io, format='PNG')
                    img_io.seek(0)

                    img = OpenpyxlImage(img_io)
                    img.anchor = f"G{row}"
                    ws.add_image(img)
                    ws.row_dimensions[row].height = 180
                except Exception:
                    # Si l'image ne peut être insérée, on laisse juste le texte
                    ws.cell(row=row, column=1, value='Preview disponible mais impossible à afficher')

                row += 5

            # Parcourir les opérations
            for op in getattr(project, "operations", []):
                # Ligne d'opération
                ws.cell(row=row, column=1, value=f"🔧 {op.label}").font = subheader_font
                ws.cell(row=row, column=2, value=op.typology or "").font = normal_font
                ws.cell(row=row, column=5, value=op.comment or "").font = normal_font
                ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

                # Calculer les temps totaux pour l'opération
                total_fixed_time = 0.0
                total_per_piece_time = 0.0
                method_comments = []

                for cost_name, cost in op.costs.items():
                    if cost.cost_type == CostType.INTERNAL_OPERATION:
                        total_fixed_time += cost.fixed_time
                        total_per_piece_time += cost.per_piece_time

                    # Collecter les commentaires méthode
                    if hasattr(cost, 'comment') and cost.comment and cost.comment.strip():
                        method_comments.append(f"{cost.name}: {cost.comment.strip()}")

                ws.cell(row=row, column=3, value=round(total_fixed_time, 3) if total_fixed_time > 0 else "").font = normal_font
                ws.cell(row=row, column=4, value=round(total_per_piece_time, 3) if total_per_piece_time > 0 else "").font = normal_font
                ws.cell(row=row, column=6, value="\n".join(method_comments)).font = normal_font
                ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

                # Appliquer les bordures
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border

                row += 1

                # Ligne de séparation si nécessaire
                if row > 2:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).border = Border(bottom=Side(style='thin'))

            # Ajuster la hauteur des lignes pour le texte wrappé
            for r in range(1, row):
                ws.row_dimensions[r].height = None  # Auto-height

            # Sauvegarder
            wb.save(output_path)

            logger.info(f"Export Fabrication/Qualité réussi → {output_path}")
            return True

        except Exception as e:
            logger.error("Erreur export Fabrication/Qualité", exc_info=True)
            raise
